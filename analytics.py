import os
import pandas as pd
import matplotlib.pyplot as plt
from app_config import get_engine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

engine = get_engine()
os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)


def save_and_report(df, chart_name, chart_type, description):
    rows = len(df)
    print(f"\n{chart_type} created ({rows} rows) → {description}")
    plt.savefig(f"charts/{chart_name}.png", bbox_inches="tight")
    plt.close()


# 1. PIE CHART — Incident types proportion per municipality
query_pie = """
SELECT r.municipality, i.type, COUNT(i.incident_id) AS total_incidents
FROM gauteng.incidents i
JOIN gauteng.roads r ON i.affected_segment_id = r.segment_id
LEFT JOIN gauteng.historical_speeds hs ON hs.segment_id = r.segment_id
GROUP BY r.municipality, i.type
ORDER BY r.municipality;
"""
df_pie = pd.read_sql(query_pie, engine)
pie_data = df_pie.groupby("type")["total_incidents"].sum()
pie_data.plot.pie(
    autopct="%1.1f%%",
    figsize=(6, 6),
    title="Distribution of Incident Types across Gauteng",
)
save_and_report(
    df_pie,
    "pie_incidents_distribution",
    "Pie Chart",
    "Proportion of incident types across all municipalities",
)


# 2. BAR CHART — Average traffic speed per road type during incidents
query_bar = """
SELECT r.road_type, i.type AS incident_type, AVG(hs.avg_speed_kph) AS avg_speed
FROM gauteng.historical_speeds hs
JOIN gauteng.roads r ON hs.segment_id = r.segment_id
JOIN gauteng.incidents i ON i.affected_segment_id = r.segment_id
GROUP BY r.road_type, i.type
ORDER BY avg_speed DESC;
"""
df_bar = pd.read_sql(query_bar, engine)
df_bar.pivot(index="road_type", columns="incident_type", values="avg_speed").plot(
    kind="bar", figsize=(8, 5)
)
plt.title("Average Speed per Road Type during Incidents")
plt.xlabel("Road Type")
plt.ylabel("Average Speed (km/h)")
plt.legend(title="Incident Type")
save_and_report(
    df_bar,
    "bar_avg_speed_incidents",
    "Bar Chart",
    "Average speed by road type and incident type",
)


# 3. HORIZONTAL BAR CHART — Top 10 roads with longest delays
query_hbar = """
SELECT r.road_type, d.priority, AVG(a.planned_duration_min) AS avg_duration
FROM gauteng.assignments a
JOIN gauteng.deliveries d ON a.delivery_id = d.delivery_id
JOIN gauteng.truck_profiles t ON d.truck_id = t.truck_id
JOIN gauteng.roads r ON r.segment_id IN (
    SELECT jsonb_array_elements_text(a.route_segments)::text
)
GROUP BY r.road_type, d.priority
ORDER BY avg_duration DESC
LIMIT 10;
"""
df_hbar = pd.read_sql(query_hbar, engine)
df_hbar.plot.barh(x="road_type", y="avg_duration", color="orange", figsize=(8, 5))
plt.title("Top Road Types by Average Delivery Duration")
plt.xlabel("Average Duration (min)")
plt.ylabel("Road Type")
save_and_report(
    df_hbar,
    "hbar_avg_duration_roads",
    "Horizontal Bar Chart",
    "Longest average delivery times by road type",
)


# 4. LINE CHART — Average daily traffic speed vs precipitation
query_line = """
SELECT DATE(hs."timestamp") AS day,
       AVG(hs.avg_speed_kph) AS avg_speed,
       AVG(w.precip_mm_h) AS avg_rain
FROM gauteng.historical_speeds hs
JOIN gauteng.weather w ON hs.segment_id = w.nearest_segment_id
LEFT JOIN gauteng.roads r ON hs.segment_id = r.segment_id
GROUP BY day
ORDER BY day;
"""
df_line = pd.read_sql(query_line, engine)
plt.figure(figsize=(8, 5))
plt.plot(df_line["day"], df_line["avg_speed"], label="Avg Speed (km/h)", marker="o")
plt.plot(df_line["day"], df_line["avg_rain"], label="Avg Rain (mm/h)", marker="x")
plt.title("Traffic Speed vs Rainfall (Daily Trend)")
plt.xlabel("Date")
plt.ylabel("Average Values")
plt.legend()
save_and_report(
    df_line,
    "line_speed_vs_rain",
    "Line Chart",
    "Shows correlation between rainfall and average speed",
)


# 5. HISTOGRAM — Distribution of delivery costs by commodity type
query_hist = """
SELECT d.commodity, d.per_km_cost_rand AS cost_per_km, t.max_weight_tons
FROM gauteng.deliveries d
JOIN gauteng.truck_profiles t ON d.truck_id = t.truck_id
LEFT JOIN gauteng.assignments a ON a.delivery_id = d.delivery_id;
"""
df_hist = pd.read_sql(query_hist, engine)
df_hist["cost_per_km"].plot.hist(bins=20, alpha=0.7)
plt.title("Distribution of Delivery Cost per Kilometer")
plt.xlabel("Cost per km (Rand)")
plt.ylabel("Frequency")
save_and_report(
    df_hist,
    "hist_cost_per_km",
    "Histogram",
    "Distribution of delivery cost per kilometer",
)


# 6. SCATTER PLOT — Relationship between delivery distance and duration
query_scatter = """
SELECT a.planned_distance_km, a.planned_duration_min, d.priority, t.max_weight_tons
FROM gauteng.assignments a
JOIN gauteng.deliveries d ON a.delivery_id = d.delivery_id
JOIN gauteng.truck_profiles t ON d.truck_id = t.truck_id
JOIN gauteng.roads r ON r.segment_id IN (SELECT jsonb_array_elements_text(a.route_segments)::text)
WHERE a.planned_distance_km IS NOT NULL AND a.planned_duration_min IS NOT NULL;
"""
df_scatter = pd.read_sql(query_scatter, engine)
plt.figure(figsize=(7, 5))
plt.scatter(
    df_scatter["planned_distance_km"],
    df_scatter["planned_duration_min"],
    c=df_scatter["max_weight_tons"],
    cmap="coolwarm",
    alpha=0.7,
)
plt.title("Delivery Distance vs Duration (by Truck Weight)")
plt.xlabel("Planned Distance (km)")
plt.ylabel("Planned Duration (min)")
plt.colorbar(label="Truck Max Weight (tons)")
save_and_report(
    df_scatter,
    "scatter_distance_vs_duration",
    "Scatter Plot",
    "Delivery distance vs duration, colored by truck capacity",
)

print("\n All charts saved in /charts/ folder.")


# Export to Excel with full formatting
def export_to_excel(dataframes_dict, filename):
    """Export multiple DataFrames to one Excel file with full formatting."""
    filepath = f"exports/{filename}"

    # Write DataFrames to Excel
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    total_rows = sum(len(df) for df in dataframes_dict.values())
    print(
        f"\n Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows"
    )

    wb = load_workbook(filepath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        ws.freeze_panes = "B2"

        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill(
            start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid"
        )
        for cell in ws[1]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, col in enumerate(ws.columns, 1):
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 3

        for col_idx, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), 1):
            if all(isinstance(c.value, (int, float)) or c.value is None for c in col):
                col_letter = get_column_letter(col_idx)
                rule = ColorScaleRule(
                    start_type="min",
                    start_color="FFAA0000",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFFFFF00",
                    end_type="max",
                    end_color="FF00AA00",
                )
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}", rule
                )

    wb.save(filepath)
    print(f" Excel formatted and saved → {filepath}")


export_to_excel(
    {
        "Incidents_by_Municipality": df_pie,
        "Speed_by_RoadType": df_bar,
        "Delivery_Costs": df_hist,
    },
    "gauteng_analytics_report.xlsx",
)
